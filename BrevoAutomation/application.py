import os
# import zipfile # Not needed for .xlsx directly
# from io import BytesIO # Not needed for saving directly to file
import openpyxl
import re
import datetime
import logging


log_filename = f"script_log_{datetime.datetime.now():%Y%m%d_%H%M%S}.log"

logging.basicConfig(
    level=logging.INFO, # Capture INFO, WARNING, ERROR, CRITICAL
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_filename), # Log to file
        logging.StreamHandler() # Also log to console (optional)
    ]
)


try:

    from BrevoAutomation.droprequester import DropRequester
except ImportError:

    try:
        from droprequester import DropRequester
    except ImportError as e:
        logging.exception(f"Error importing DropRequester: {e}")
        logging.info("Make sure droprequester.py is accessible (e.g., in the same directory or BrevoAutomation is a package)")
        exit(1)



#Import BrevoRequester

try:

    from BrevoAutomation.brevorequester import BrevoRequester
except ImportError:

    try:
        from brevorequester import BrevoRequester
    except ImportError as e:
        logging.exception(f"Error importing BrevoRequester: {e}")
        logging.info("Make sure brevorequester.py is accessible (e.g., in the same directory or BrevoAutomation is a package)")
        exit(1)


if __name__ == '__main__':
    drop_requester = DropRequester()
    brevo_requester = BrevoRequester()

    # Path to the file on Dropbox
    DROPBOX_FILE_PATH = "/REO/Medlems-liste/Medlemmer-alle-01.01.2025.xlsx"
    # Local path to save the file
    LOCAL_FILENAME = "AlleMedlemmer.xlsx"
    TARGET_BREVO_LIST_ID = 4  # The "Alle Medlemmer" list ID

    try:
        # Call the download method
        download_response = drop_requester.download_file(DROPBOX_FILE_PATH)

        # --- Debugging: Print type and keys if it's a dict ---
        logging.info(f"Type of download_response: {type(download_response)}")
        if isinstance(download_response, dict):
            logging.info(f"Keys in download_response dictionary: {download_response.keys()}")
        # --- End Debugging ---


        # Check if the response is a dictionary and contains the expected keys
        file_bytes = None
        if isinstance(download_response, dict):
            if 'content' in download_response:
                 file_bytes = download_response['content']
            elif 'data' in download_response:
                 file_bytes = download_response['data']
            # Add other potential key checks if needed
        elif isinstance(download_response, bytes):
            # save directly if it's bytes
            file_bytes = download_response
        elif hasattr(download_response, 'content'):
             # if the response is an object with a content attribute, save that.
             file_bytes = download_response.content

        # Check if we successfully extracted bytes
        if file_bytes is None:
            logging.error("Error: Could not find file content in the download response.")
            logging.info(f"Response received: {download_response}")
            exit(1)
        elif not isinstance(file_bytes, bytes):
             logging.error(f"Error: Expected bytes for file content, but found type {type(file_bytes)}.")
             logging.info(f"Content value (first 100 bytes): {file_bytes[:100]}")
             exit(1)

        logging.info("File content bytes extracted successfully.")

        # Save the extracted bytes to a local file
        with open(LOCAL_FILENAME, "wb") as f:
            f.write(file_bytes)

        logging.info(f"File saved successfully as {LOCAL_FILENAME}.")

    except FileNotFoundError:
        logging.exception(f"Error: The file was not found on Dropbox at path: {DROPBOX_FILE_PATH}")
    except Exception as e:
        logging.exception(f"An error occurred during download or processing: {e}")



    def prepare_brevo_contacts_from_excel(filename="AlleMedlemmer.xlsx", target_list_id=4):
        """
        Reads an Excel file, extracts Email and First Name, and formats
        data for the Brevo 'create_contact' API endpoint.

        Args:
            filename (str): The path to the Excel file.
            target_list_id (int): The Brevo list ID to add contacts to.

        Returns:
            list: A list of dictionaries, each formatted for Brevo's
                  create_contact API. Returns an empty list if errors occur
                  or no valid contacts are found.
        """
        brevo_contacts = []
        try:
            workbook = openpyxl.load_workbook(filename=filename,
                                              data_only=True)  # data_only for cell values, not formulas
            sheet = workbook.active
        except FileNotFoundError:
            logging.error(f"Error: Excel file not found at {filename}")
            return []
        except Exception as e:
            logging.exception(f"Error loading Excel workbook {filename}: {e}")
            return []

        header_row_index = -1
        email_col_index = -1
        fname_col_index = -1

        # Find header row and column indices (more robust than assuming row 2)
        # Look for 'Mail' and 'ForNavn' in the first few rows
        MAX_HEADER_SEARCH_ROWS = 5
        for r_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=MAX_HEADER_SEARCH_ROWS, values_only=True)):
            # Convert potential None values in header to empty strings for safe checking
            row_values = [str(val).strip() if val is not None else "" for val in row]
            try:
                # Case-insensitive search for headers
                lower_row_values = [val.lower() for val in row_values]
                if 'mail' in lower_row_values and 'fornavn' in lower_row_values:
                    email_col_index = lower_row_values.index('mail')
                    fname_col_index = lower_row_values.index('fornavn')
                    header_row_index = r_idx + 1  # 1-based index for openpyxl
                    logging.info(f"Header row found at index {header_row_index}.")
                    logging.info(f"Email column index: {email_col_index}, First Name column index: {fname_col_index}")
                    break
            except ValueError:
                # Column not found in this row, continue searching
                continue

        if header_row_index == -1:
            logging.error(
                f"Error: Could not find header row with 'Mail' and 'ForNavn' within the first {MAX_HEADER_SEARCH_ROWS} rows.")
            return []

        # Basic email format check regex
        email_regex = r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$"

        processed_emails = set()  # Use a set to avoid duplicate emails

        # Iterate through each data row after the header.
        for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row_index + 1, values_only=True),
                                      start=header_row_index + 1):
            try:
                email_raw = row[email_col_index]
                fname_raw = row[fname_col_index]

                # Data Cleaning & Validation
                email = str(email_raw).strip() if email_raw else None
                fname = str(fname_raw).strip() if fname_raw else None

                # Basic validation
                if not email or not re.match(email_regex, email):
                    logging.info(f"Skipping row {row_idx}: Invalid or missing email ('{email_raw}')")
                    continue

                if email.lower() in processed_emails:
                    logging.info(f"Skipping row {row_idx}: Duplicate email in file ('{email}')")
                    continue

                # If first name is missing, default to something or skip, depending on policy
                # Brevo might require a non-empty value if an attribute is included.
                # Using an empty string might be acceptable. Using the email part might be another option.
                if not fname:
                    fname = ""  # Or potentially skip the contact if First Name is mandatory
                    logging.warning(f"Warning: Missing First Name for email '{email}' in row {row_idx}. Setting to empty string.")

                # --- Format for Brevo ---
                contact_data = {
                    "email": email,
                    "attributes": {
                        # Use Brevo's standard uppercase attribute names unless customized otherwise
                        "FIRSTNAME": fname
                        # Add LASTNAME here if needed later: "LASTNAME": lname
                    },
                    "listIds": [target_list_id],
                    "updateEnabled": True  # Update contact if email exists. THIS IS VERY IMPORTANT
                }
                brevo_contacts.append(contact_data)
                processed_emails.add(email.lower())

            except IndexError:
                logging.warning(f"Skipping row {row_idx}: Row has fewer columns than expected.")
                continue
            except Exception as e:
                logging.exception(f"Skipping row {row_idx}: Unexpected error processing row - {e}")
                continue

        logging.info(f"Prepared {len(brevo_contacts)} unique contacts for Brevo.")
        return brevo_contacts




    logging.info(f"\n--- Preparing contacts from Excel file: {LOCAL_FILENAME} ---")
    try:
        contacts_for_brevo = prepare_brevo_contacts_from_excel(
            filename=LOCAL_FILENAME,
            target_list_id=TARGET_BREVO_LIST_ID
        )
        if not contacts_for_brevo:
            logging.error("No valid contacts were prepared from the Excel file. Exiting.")
            # Optional: Clean up downloaded file
            if os.path.exists(LOCAL_FILENAME):
                os.remove(LOCAL_FILENAME)
            exit(0) # Exit gracefully if no contacts found

    except Exception as e:
        logging.exception(f"An error occurred during Excel processing: {e}")
        exit(1)


    # Update Contacts in Brevo
    try:
        brevo_requester.update_all_members_contact_list(
            contacts_to_process=contacts_for_brevo,
            target_list_id=TARGET_BREVO_LIST_ID
        )
    except Exception as e:
        logging.exception(f"An error occurred during Brevo update: {e}")

        exit(1)